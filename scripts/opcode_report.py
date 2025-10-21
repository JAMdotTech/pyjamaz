#!/usr/bin/env python3
"""Pretty-print PVM opcode timing stats stored in the SQLite database."""

from __future__ import annotations

import argparse
import sqlite3
import sys
from pathlib import Path
from typing import Dict, Iterable, Optional, Sequence, Tuple

try:
    from openpyxl import Workbook
except ImportError:  # pragma: no cover - soft dependency
    Workbook = None


DEFAULT_DB_PATH = Path.cwd() / "opcode_timing_stats.sqlite3"
DEFAULT_XLSX_PATH = Path.cwd() / "opcode_timing_report.xlsx"


RunRow = Tuple[int, str, str, int, float, float]
SampleRow = Tuple[int, str, int, float, float, float, float, float]


def parse_args(argv: Sequence[str]) -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Read opcode timing stats from SQLite and print a formatted report."
    )
    parser.add_argument(
        "--db",
        type=Path,
        default=DEFAULT_DB_PATH,
        help=f"Path to opcode timing SQLite database (default: {DEFAULT_DB_PATH})",
    )
    parser.add_argument(
        "--limit",
        type=int,
        default=100,
        help=(
            "Maximum number of runs to display when using --per-run "
            "(ignored when --run-id is supplied)."
        ),
    )
    parser.add_argument(
        "--run-id",
        type=int,
        help="Show a specific run id instead of the most recent runs.",
    )
    parser.add_argument(
        "--source",
        type=str,
        help="Filter runs by source label (exact match).",
    )
    parser.add_argument(
        "--excel",
        type=Path,
        default=DEFAULT_XLSX_PATH,
        help=f"Path for Excel export (default: {DEFAULT_XLSX_PATH}).",
    )
    parser.add_argument(
        "--no-excel",
        action="store_true",
        help="Skip Excel export.",
    )
    parser.add_argument(
        "--per-run",
        action="store_true",
        help="Report stats for individual runs instead of aggregating all runs.",
    )
    return parser.parse_args(argv)


def fetch_runs(
    conn: sqlite3.Connection,
    limit: Optional[int],
    run_id: Optional[int],
    source: Optional[str],
) -> Iterable[RunRow]:
    base_query = (
        "SELECT id, created_at, source, total_iterations, total_time, avg_time_per_instr "
        "FROM opcode_timing_runs"
    )
    params: list = []
    clauses = []
    if run_id is not None:
        clauses.append("id = ?")
        params.append(run_id)
    if source:
        clauses.append("source = ?")
        params.append(source)

    if clauses:
        base_query += " WHERE " + " AND ".join(clauses)

    base_query += " ORDER BY id DESC"

    if run_id is None and limit is not None:
        base_query += " LIMIT ?"
        params.append(limit)

    cur = conn.execute(base_query, params)
    return list(cur.fetchall())


def fetch_samples(conn: sqlite3.Connection, run_id: int) -> Iterable[SampleRow]:
    cur = conn.execute(
        """
        SELECT opcode, opcode_name, count, total_time, avg_time, mean_time, min_time, max_time
        FROM opcode_timing_samples
        WHERE run_id = ?
        ORDER BY mean_time DESC, max_time DESC, opcode ASC
        """,
        (run_id,),
    )
    return list(cur.fetchall())


def format_run(run: RunRow, samples: Iterable[SampleRow]) -> str:
    run_id, created_at, source, total_iterations, total_time, avg_time_per_instr = run
    lines = []
    header = (
        f"Run #{run_id} @ {created_at} [source={source}]"
        if source
        else f"Run #{run_id} @ {created_at}"
    )
    lines.append(header)
    lines.append(f"Opcode timing stats: total iterations {total_iterations}")

    if total_iterations == 0:
        lines.append("Opcode timing stats: no iterations executed")
        return "\n".join(lines)

    lines.append(f"Total recorded time (seconds) {total_time:.9f}")
    lines.append(
        "Average time per instruction (seconds) "
        f"{avg_time_per_instr:.9e} ({avg_time_per_instr * 1e6:.3f} us)"
    )

    for opcode, name, count, total_time, avg_time, mean_time, min_time, max_time in samples:
        lines.append(
            f"   {opcode:3d} {name:>24} "
            f"count {count:6d} total_us {total_time * 1e6:.3f} "
            f"avg_us {avg_time * 1e6:.3f} mean_us {mean_time * 1e6:.3f} "
            f"min_us {min_time * 1e6:.3f} max_us {max_time * 1e6:.3f}"
        )

    return "\n".join(lines)


def export_to_excel(
    runs: Iterable[RunRow],
    samples_by_run: Dict[int, Iterable[SampleRow]],
    path: Path,
) -> None:
    path = path.resolve()
    path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    # Remove default sheet to keep only our generated ones
    default_ws = wb.active
    wb.remove(default_ws)

    for run in runs:
        run_id, created_at, source, total_iterations, total_time, avg_time_per_instr = run
        ws = wb.create_sheet(f"run_{run_id}")

        metadata_rows = [
            ("run_id", run_id),
            ("created_at", created_at),
            ("source", source),
            ("total_iterations", total_iterations),
            ("total_time_s", total_time),
            ("avg_time_per_instr_s", avg_time_per_instr),
            ("avg_time_per_instr_us", avg_time_per_instr * 1e6),
        ]

        for label, value in metadata_rows:
            ws.append([label, value])

        ws.append([])  # Blank separator row

        header = [
            "opcode",
            "opcode_name",
            "count",
            "total_time_s",
            "total_time_us",
            "avg_time_s",
            "avg_time_us",
            "mean_time_s",
            "mean_time_us",
            "min_time_s",
            "min_time_us",
            "max_time_s",
            "max_time_us",
        ]
        ws.append(header)

        for sample in samples_by_run.get(run_id, []):
            (
                opcode,
                name,
                count,
                total_time_sample,
                avg_time,
                mean_time,
                min_time,
                max_time,
            ) = sample

            ws.append([
                opcode,
                name,
                count,
                total_time_sample,
                total_time_sample * 1e6,
                avg_time,
                avg_time * 1e6,
                mean_time,
                mean_time * 1e6,
                min_time,
                min_time * 1e6,
                max_time,
                max_time * 1e6,
            ])

    wb.save(path)
    print(f"Excel report written to {path}")


def fetch_aggregate_samples(
    conn: sqlite3.Connection,
    run_ids: Sequence[int],
) -> Iterable[Tuple[int, str, int, float, float, float, float, float]]:
    if not run_ids:
        return []

    placeholders = ",".join("?" for _ in run_ids)
    query = f"""
        SELECT
            opcode,
            opcode_name,
            SUM(count) AS total_count,
            SUM(total_time) AS total_time,
            CASE WHEN SUM(count) = 0 THEN 0.0
                 ELSE SUM(total_time) / SUM(count)
            END AS avg_time,
            CASE WHEN SUM(count) = 0 THEN 0.0
                 ELSE SUM(mean_time * count) / SUM(count)
            END AS mean_time,
            MIN(min_time) AS min_time,
            MAX(max_time) AS max_time
        FROM opcode_timing_samples
        WHERE run_id IN ({placeholders})
        GROUP BY opcode, opcode_name
        ORDER BY avg_time DESC, max_time DESC, opcode ASC
    """
    cur = conn.execute(query, list(run_ids))
    return list(cur.fetchall())


def format_aggregate(runs: Sequence[RunRow], samples: Iterable[SampleRow]) -> str:
    run_count = len(runs)
    total_iterations = sum(run[3] for run in runs)
    total_time = sum(run[4] for run in runs)
    avg_time_per_instr = (
        total_time / total_iterations if total_iterations else 0.0
    )

    lines = [f"Aggregated opcode timing stats across {run_count} run(s)"]

    lines.append(f"Total iterations {total_iterations}")

    if total_iterations == 0:
        lines.append("Opcode timing stats: no iterations executed")
        return "\n".join(lines)

    lines.append(f"Total recorded time (seconds) {total_time:.9f}")
    lines.append(
        "Average time per instruction (seconds) "
        f"{avg_time_per_instr:.9e} ({avg_time_per_instr * 1e6:.3f} us)"
    )

    for opcode, name, count, total_time_sample, avg_time, mean_time, min_time, max_time in samples:
        lines.append(
            f"   {opcode:3d} {name:>24} "
            f"count {count:6d} total_us {total_time_sample * 1e6:.3f} "
            f"avg_us {avg_time * 1e6:.3f} mean_us {mean_time * 1e6:.3f} "
            f"min_us {min_time * 1e6:.3f} max_us {max_time * 1e6:.3f}"
        )

    return "\n".join(lines)


def export_aggregate_to_excel(
    runs: Sequence[RunRow],
    samples: Iterable[SampleRow],
    path: Path,
) -> None:
    path = path.resolve()
    path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    default_ws = wb.active
    wb.remove(default_ws)

    run_count = len(runs)
    total_iterations = sum(run[3] for run in runs)
    total_time = sum(run[4] for run in runs)
    avg_time_per_instr = (
        total_time / total_iterations if total_iterations else 0.0
    )

    summary_ws = wb.create_sheet("aggregate")
    summary_rows = [
        ("run_count", run_count),
        ("total_iterations", total_iterations),
        ("total_time_s", total_time),
        ("avg_time_per_instr_s", avg_time_per_instr),
        ("avg_time_per_instr_us", avg_time_per_instr * 1e6),
    ]
    for label, value in summary_rows:
        summary_ws.append([label, value])

    summary_ws.append([])

    header = [
        "opcode",
        "opcode_name",
        "count",
        "total_time_s",
        "total_time_us",
        "avg_time_s",
        "avg_time_us",
        "mean_time_s",
        "mean_time_us",
        "min_time_s",
        "min_time_us",
        "max_time_s",
        "max_time_us",
    ]
    summary_ws.append(header)

    for sample in samples:
        (
            opcode,
            name,
            count,
            total_time_sample,
            avg_time,
            mean_time,
            min_time,
            max_time,
        ) = sample

        summary_ws.append([
            opcode,
            name,
            count,
            total_time_sample,
            total_time_sample * 1e6,
            avg_time,
            avg_time * 1e6,
            mean_time,
            mean_time * 1e6,
            min_time,
            min_time * 1e6,
            max_time,
            max_time * 1e6,
        ])

    runs_ws = wb.create_sheet("included_runs")
    runs_ws.append(
        [
            "run_id",
            "created_at",
            "source",
            "total_iterations",
            "total_time_s",
            "avg_time_per_instr_s",
            "avg_time_per_instr_us",
        ]
    )
    for run in runs:
        run_id, created_at, source, total_iterations_run, total_time_run, avg_time_run = run
        runs_ws.append(
            [
                run_id,
                created_at,
                source,
                total_iterations_run,
                total_time_run,
                avg_time_run,
                avg_time_run * 1e6,
            ]
        )

    wb.save(path)
    print(f"Excel report written to {path}")


def main(argv: Sequence[str]) -> int:
    args = parse_args(argv)

    if not args.db.exists():
        print(f"Database not found: {args.db}", file=sys.stderr)
        return 1

    conn = sqlite3.connect(str(args.db))
    conn.execute("PRAGMA foreign_keys = ON")

    try:
        if args.per_run:
            runs = fetch_runs(conn, args.limit, args.run_id, args.source)

            if not runs:
                target = f"run_id={args.run_id}" if args.run_id else "recent runs"
                print(f"No opcode timing data found for {target}.")
                return 0

            rendered = []
            samples_by_run: Dict[int, Iterable[SampleRow]] = {}
            for run in runs:
                samples = fetch_samples(conn, run[0])
                samples_by_run[run[0]] = samples
                rendered.append(format_run(run, samples))

            print("\n\n".join(rendered))

            excel_path: Optional[Path] = None if args.no_excel else args.excel
            if excel_path:
                if Workbook is None:
                    print(
                        "[opcode_timing_report] openpyxl is required for Excel export. "
                        "Install it or re-run with --no-excel.",
                        file=sys.stderr,
                    )
                else:
                    export_to_excel(runs, samples_by_run, excel_path)
        else:
            runs = fetch_runs(conn, None, args.run_id, args.source)

            if not runs:
                if args.run_id:
                    print(f"No opcode timing data found for run_id={args.run_id}.")
                elif args.source:
                    print(f"No opcode timing data found for source={args.source}.")
                else:
                    print("No opcode timing data found.")
                return 0

            run_ids = [run[0] for run in runs]
            aggregate_samples = fetch_aggregate_samples(conn, run_ids)
            print(format_aggregate(runs, aggregate_samples))

            excel_path = None if args.no_excel else args.excel
            if excel_path:
                if Workbook is None:
                    print(
                        "[opcode_timing_report] openpyxl is required for Excel export. "
                        "Install it or re-run with --no-excel.",
                        file=sys.stderr,
                    )
                else:
                    export_aggregate_to_excel(runs, aggregate_samples, excel_path)
        return 0
    finally:
        conn.close()


if __name__ == "__main__":
    raise SystemExit(main(sys.argv[1:]))

